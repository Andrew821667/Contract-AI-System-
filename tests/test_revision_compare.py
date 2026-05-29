# -*- coding: utf-8 -*-
"""
Unit tests for the revision-compare stack:
  - RevisionComparator (matching + heuristic classification)
  - revision_xlsx_exporter (template-matched workbook layout)
  - revision_pdf_exporter (landscape A3 PDF using reportlab)

Real LLM and DB are *not* exercised here. The LLM path is implicitly
covered when llm_gateway is None, which forces the deterministic
heuristic — exactly what the API endpoint falls back to when the
gateway is offline.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.services.revision_comparator import (
    Assessment,
    Perspective,
    RevisionComparator,
    RevisionDiffReport,
    RiskLevel,
)
from src.services.revision_pdf_exporter import export_report as pdf_export_report
from src.services.revision_xlsx_exporter import export_report as xlsx_export_report


# --- fixtures -------------------------------------------------------------

OLD_CLAUSES = [
    {"number": "1.1", "title": "Предмет / терминология",
     "text": "Товар/Зерно: плоды злаковых, зернобобовых и масличных культур."},
    {"number": "2.1", "title": "Базис поставки",
     "text": "Поставка осуществляется в адрес Покупателя или Грузополучателя."},
    {"number": "2.3", "title": "Оплата",
     "text": "Покупатель вправе продлить оплату до предоставления документов."},
    {"number": "2.8", "title": "Хранение",
     "text": "Подробный режим хранения у Покупателя до перехода собственности."},
]
NEW_CLAUSES = [
    {"number": "1.1", "title": "Предмет / терминология",
     "text": "Товар: плоды злаковых, зернобобовых и масличных культур."},
    {"number": "2.1", "title": "Базис поставки",
     "text": "Поставка на условиях EXW — склад Поставщика по INCOTERMS-2010."},
    {"number": "2.4", "title": "Оплата",
     "text": "Покупатель вправе продлить оплату до предоставления документов."},
    {"number": "2.2", "title": "Количество",
     "text": "Допускается отклонение ±10% в пределах одного транспортного средства."},
]


class _SwitchingParser:
    """Returns OLD_CLAUSES on first call, NEW_CLAUSES on every later call."""

    def __init__(self) -> None:
        self._calls = 0

    def extract_clauses(self, _content: str) -> list[dict]:
        self._calls += 1
        return OLD_CLAUSES if self._calls == 1 else NEW_CLAUSES


@pytest.fixture
def heuristic_report() -> RevisionDiffReport:
    comp = RevisionComparator(
        parser=_SwitchingParser(),
        old_revision_label="Редакция 2025",
        new_revision_label="Редакция 2026",
    )
    return comp.compare(
        "OLD_CONTENT", "NEW_CONTENT",
        perspective=Perspective.SUPPLIER,
        title="Сравнение редакций — тестовый отчёт",
        old_file_name="old.pdf",
        new_file_name="new.pdf",
    )


# --- comparator -----------------------------------------------------------

def test_comparator_matches_clauses_by_number(heuristic_report: RevisionDiffReport) -> None:
    # Three clauses share the same number across revisions (1.1, 2.1, 2.4 is
    # only in new but matches no old, 2.8 only old). Make sure we emit one
    # row per logical clause and label it accordingly.
    labels = [r.clause_pair_label for r in heuristic_report.rows]

    # Matched clauses
    assert "п.1.1 ↔ п.1.1" in labels
    assert "п.2.1 ↔ п.2.1" in labels

    # Deleted clauses
    assert any("(удалён)" in lbl for lbl in labels)

    # Added clauses
    assert any("(новый)" in lbl for lbl in labels)


def test_comparator_perspective_propagates(heuristic_report: RevisionDiffReport) -> None:
    assert heuristic_report.perspective is Perspective.SUPPLIER
    assert "Поставщик" in heuristic_report.summary.overall_verdict


def test_comparator_rows_have_all_fields(heuristic_report: RevisionDiffReport) -> None:
    """Every row produced by the heuristic must have non-empty values for
    every column we render in xlsx/pdf — otherwise exporters silently
    drop content."""
    for row in heuristic_report.rows:
        assert row.number > 0
        assert row.clause_pair_label
        assert row.block
        assert row.condition
        assert row.change_summary
        assert isinstance(row.assessment, Assessment)
        assert isinstance(row.risk_level, RiskLevel)
        assert row.complex_impact
        assert row.recommendation
        assert row.source


# --- _ParserAdapter (routes.py) -------------------------------------------

def test_parser_adapter_handles_xml_envelope_from_parse_txt(tmp_path: Path) -> None:
    """Regression: comparator must extract clauses even when DocumentParser
    wraps the source text in its <contract>...<clauses><clause> XML
    envelope (which happens for .txt and .pdf input). Before the fix the
    adapter ran the regex on raw XML and produced one giant blob."""
    from src.api.revisions.routes import _ParserAdapter
    from src.services.document_parser import DocumentParser

    src = tmp_path / "old.txt"
    src.write_text(
        "1.1 Предмет договора\nПоставка зерна урожая 2025 года.\n"
        "2.1 Базис поставки\nПоставка в адрес Покупателя или Грузополучателя.\n"
        "2.3 Условия оплаты\nПокупатель оплачивает в течение 30 банковских дней.\n",
        encoding="utf-8",
    )
    adapter = _ParserAdapter(DocumentParser())
    xml_content = adapter.parse(str(src))

    clauses = adapter.extract_clauses(xml_content)
    numbers = [c.get("number") for c in clauses if c.get("number")]
    # At least three numbered clauses must come out (1.1, 2.1, 2.3) — not
    # a single XML-blob row.
    assert len(numbers) >= 3, f"expected ≥3 numbered clauses, got: {clauses}"
    assert "1.1" in numbers
    assert "2.1" in numbers


# --- xlsx exporter --------------------------------------------------------

def test_xlsx_exporter_layout(tmp_path: Path, heuristic_report: RevisionDiffReport) -> None:
    out = tmp_path / "report.xlsx"
    xlsx_export_report(heuristic_report, out,
                       old_revision_label="Редакция 2025",
                       new_revision_label="Редакция 2026")

    wb = load_workbook(out)
    assert wb.sheetnames == ["Сравнение условий", "Краткие выводы"]

    diff = wb["Сравнение условий"]
    # 12 columns (10 from the original template + 'Пункт' + 'Комплексное влияние')
    assert diff.max_column == 12
    # 1 header row + N data rows
    assert diff.max_row == 1 + len(heuristic_report.rows)

    expected_headers = [
        "№", "Пункт", "Блок", "Условие",
        "Редакция 2025", "Редакция 2026", "Изменение / несоответствие",
        "Оценка для «Поставщик»", "Риск",
        "Комплексное влияние на договор", "Рекомендация", "Источник",
    ]
    actual_headers = [c.value for c in diff[1]]
    assert actual_headers == expected_headers

    # Header styling: dark-blue fill + bold white Arial 10pt
    cell = diff["A1"]
    assert cell.font.bold is True
    assert cell.font.name == "Arial"
    assert int(cell.font.size) == 10
    assert (cell.fill.fgColor.rgb or "").upper().endswith("1F4E78")


def test_xlsx_summary_sheet_has_title_and_label_column(
    tmp_path: Path, heuristic_report: RevisionDiffReport,
) -> None:
    out = tmp_path / "report.xlsx"
    xlsx_export_report(heuristic_report, out)
    wb = load_workbook(out)
    summary = wb["Краткие выводы"]

    # Row 1 is the merged title (same dark-blue header treatment as diff sheet)
    title_cell = summary["A1"]
    assert title_cell.value == "Сравнение редакций — тестовый отчёт"
    assert title_cell.font.bold is True

    # Row 3 is the first label row ("Дата подготовки") with light-blue fill
    label = summary["A3"]
    assert label.value == "Дата подготовки"
    assert (label.fill.fgColor.rgb or "").upper().endswith("D9EAF7")


# --- pdf exporter ---------------------------------------------------------

def test_pdf_exporter_produces_nonempty_file(
    tmp_path: Path, heuristic_report: RevisionDiffReport,
) -> None:
    out = tmp_path / "report.pdf"
    pdf_export_report(heuristic_report, out,
                      old_revision_label="Редакция 2025",
                      new_revision_label="Редакция 2026")
    assert out.is_file()
    # A minimal landscape-A3 PDF with a title block + table has to be at
    # least a few KB. Anything smaller means content didn't render.
    assert out.stat().st_size > 2 * 1024
    # PDF magic bytes — smoke check.
    assert out.read_bytes()[:5] == b"%PDF-"
