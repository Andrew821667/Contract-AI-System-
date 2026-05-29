# -*- coding: utf-8 -*-
"""
Excel exporter for RevisionDiffReport.

Produces a two-sheet xlsx that mirrors the layout used by the lawyer
team (см. образец 'Сравнение_редакций_договора_..._доп_риски.xlsx'):

  Sheet 1 «Сравнение условий»: №, Блок, Условие, <Старая редакция>,
    <Новая редакция>, Изменение / несоответствие, Оценка для <стороны>,
    Риск, Рекомендация, Источник.
  Sheet 2 «Краткие выводы»: ключ → значение, с разделами заголовков,
    общий вывод, ключевые плюсы / риски, что править перед подписанием.

Column widths and header fill match the template exactly. Cells in the
"Оценка" and "Риск" columns are colour-coded so the report is scannable.
"""
from __future__ import annotations

from pathlib import Path
from typing import Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .revision_comparator import (
    ASSESSMENT_LABELS_RU,
    PERSPECTIVE_LABELS_RU,
    RISK_LABELS_RU,
    RevisionDiffReport,
    RevisionDiffRow,
)


# Match the existing template exactly so reports from this code drop
# straight into the lawyer team's workflow.
#
# Everything here is copied 1:1 from the reference workbook the lawyer
# team uses (header fill, header text colour, font family/size, row
# heights, alignment, no borders, no data-row fills). The only two
# additions we make on top of the original 10 columns are:
#   - column "Пункт"            (col B) — clause-pair label
#   - column "Комплексное влияние" (col J) — impact on the rest of the deal
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)

# Summary-sheet label background — matches the existing template (light blue).
SUMMARY_LABEL_FILL = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
SUMMARY_LABEL_FONT = Font(name="Arial", bold=True, size=10)


SHEET_DIFF = "Сравнение условий"
SHEET_SUMMARY = "Краткие выводы"

# Column widths. Original template uses (5, 20, 32, 38, 38, 38, 16, 16, 44, 24).
# We keep those for matching columns and slot in two new ones at B (Пункт)
# and J (Комплексное влияние).
#                        A   B   C   D   E   F   G   H   I   J   K   L
DIFF_COL_WIDTHS =       (5, 16, 20, 32, 38, 38, 38, 16, 16, 38, 44, 24)
DIFF_HEADER_ROW_HEIGHT = 34
DIFF_DATA_ROW_HEIGHT   = 75


def export_report(
    report: RevisionDiffReport,
    output_path: Union[str, Path],
    *,
    old_revision_label: str = "Редакция 1",
    new_revision_label: str = "Редакция 2",
) -> Path:
    """Render `report` into an xlsx at `output_path`.

    Returns the resolved Path. Overwrites if the file already exists.
    """
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    diff_ws = wb.active
    diff_ws.title = SHEET_DIFF
    _write_diff_sheet(diff_ws, report, old_revision_label, new_revision_label)

    summary_ws = wb.create_sheet(SHEET_SUMMARY)
    _write_summary_sheet(summary_ws, report)

    wb.save(out)
    return out


# --- diff sheet -----------------------------------------------------------

def _write_diff_sheet(
    ws: Worksheet,
    report: RevisionDiffReport,
    old_label: str,
    new_label: str,
) -> None:
    perspective_label = PERSPECTIVE_LABELS_RU[report.perspective]
    headers = (
        "№",
        "Пункт",
        "Блок",
        "Условие",
        old_label,
        new_label,
        "Изменение / несоответствие",
        f"Оценка для «{perspective_label}»",
        "Риск",
        "Комплексное влияние на договор",
        "Рекомендация",
        "Источник",
    )

    # Header row — same fill/font/alignment as the template
    for col_idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    for col_idx, width in enumerate(DIFF_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = DIFF_HEADER_ROW_HEIGHT
    # Keep header visible while scrolling — small extra over the original
    # template (which didn't freeze), but harmless and useful for long reports.
    ws.freeze_panes = "A2"

    # Body rows
    for row_idx, row in enumerate(report.rows, start=2):
        _write_row(ws, row_idx, row)
        ws.row_dimensions[row_idx].height = DIFF_DATA_ROW_HEIGHT


def _write_row(ws: Worksheet, row_idx: int, row: RevisionDiffRow) -> None:
    values = (
        row.number,                              # 1: №
        row.clause_pair_label,                   # 2: Пункт ("п.5.2 ↔ п.5.3")
        row.block,                               # 3: Блок
        row.condition,                           # 4: Условие
        row.old_text or "—",                     # 5: Редакция X
        row.new_text or "—",                     # 6: Редакция Y
        row.change_summary,                      # 7: Изменение / несоответствие
        ASSESSMENT_LABELS_RU[row.assessment],    # 8: Оценка
        RISK_LABELS_RU[row.risk_level],          # 9: Риск
        row.complex_impact,                      # 10: Комплексное влияние ← NEW
        row.recommendation,                      # 11: Рекомендация
        row.source,                              # 12: Источник
    )
    # №, Пункт, Оценка, Риск — центр; остальное — текстовый блок слева.
    center_cols = {1, 2, 8, 9}
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = BODY_FONT
        cell.alignment = Alignment(
            horizontal="center" if col_idx in center_cols else "left",
            vertical="top",
            wrap_text=True,
        )


# --- summary sheet --------------------------------------------------------

def _write_summary_sheet(ws: Worksheet, report: RevisionDiffReport) -> None:
    # Column widths mimic the reference workbook (narrow label column, wide
    # value column). Helps long verdicts wrap naturally.
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 90

    summary = report.summary

    # Title row: same dark-blue header treatment as the diff sheet, merged
    # across both columns. Centred, bold, white text.
    title_cell = ws.cell(row=1, column=1, value=summary.title)
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws.row_dimensions[1].height = DIFF_HEADER_ROW_HEIGHT

    rows: list[tuple[str, str]] = [
        ("Дата подготовки", summary.prepared_at.strftime("%d.%m.%Y %H:%M UTC")),
        ("Точка зрения", PERSPECTIVE_LABELS_RU[report.perspective]),
        ("Сравниваемые документы", summary.documents_compared),
        ("Общий вывод", summary.overall_verdict),
        ("Ключевые плюсы", _bulleted(summary.key_pros)),
        ("Ключевые риски", _bulleted(summary.key_risks)),
        ("Что править перед подписанием", _bulleted(summary.pre_signature_edits)),
    ]
    for label, value in summary.source_files.items():
        if value:
            rows.append((label, value))

    for r_offset, (label, value) in enumerate(rows, start=3):
        label_cell = ws.cell(row=r_offset, column=1, value=label)
        label_cell.font = SUMMARY_LABEL_FONT
        label_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        label_cell.fill = SUMMARY_LABEL_FILL

        val_cell = ws.cell(row=r_offset, column=2, value=value)
        val_cell.font = BODY_FONT
        val_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Tall rows for multi-line bullet lists so the text isn't clipped.
        if "\n" in (value or ""):
            ws.row_dimensions[r_offset].height = max(60, value.count("\n") * 18)
        else:
            ws.row_dimensions[r_offset].height = 30


def _bulleted(items: list[str]) -> str:
    if not items:
        return "—"
    return "\n".join(f"• {item}" for item in items)


__all__ = ["export_report"]
